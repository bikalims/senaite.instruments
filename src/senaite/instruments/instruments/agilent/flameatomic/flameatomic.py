# -*- coding: utf-8 -*-
#
# This file is part of SENAITE.INSTRUMENTS.
#
# SENAITE.CORE is free software: you can redistribute it and/or modify it under
# the terms of the GNU General Public License as published by the Free Software
# Foundation, version 2.
#
# This program is distributed in the hope that it will be useful, but WITHOUT
# ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS
# FOR A PARTICULAR PURPOSE. See the GNU General Public License for more
# details.
#
# You should have received a copy of the GNU General Public License along with
# this program; if not, write to the Free Software Foundation, Inc., 51
# Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.
#
# Copyright 2018-2019 by it's authors.
# Some rights reserved, see README and LICENSE.

import re
import json
import types
import traceback
from cStringIO import StringIO
from DateTime import DateTime
from mimetypes import guess_type
from openpyxl import load_workbook
from os.path import abspath
from os.path import splitext
from xlrd import open_workbook
from bika.lims.browser import BrowserView

from senaite.core.exportimport.instruments import (
    IInstrumentAutoImportInterface, IInstrumentImportInterface
)
from senaite.core.exportimport.instruments.resultsimport import (
    AnalysisResultsImporter)
from senaite.core.exportimport.instruments.resultsimport import (
    InstrumentResultsFileParser)
from senaite.core.exportimport.instruments import IInstrumentExportInterface

from bika.lims import api
from bika.lims import bikaMessageFactory as _
from bika.lims.catalog import CATALOG_ANALYSIS_REQUEST_LISTING
from senaite.core.catalog import ANALYSIS_CATALOG, SENAITE_CATALOG
from senaite.instruments.instrument import FileStub
from senaite.instruments.instrument import SheetNotFound
from zope.interface import implements
from zope.publisher.browser import FileUpload
from zope.component import getUtility
from plone.i18n.normalizer.interfaces import IIDNormalizer


class SampleNotFound(Exception):
    pass


class MultipleAnalysesFound(Exception):
    pass


class AnalysisNotFound(Exception):
    pass


class FlameAtomicParser(InstrumentResultsFileParser):
    ar = None

    def __init__(self, infile, worksheet=None, encoding=None, delimiter=None):
        self.delimiter = delimiter if delimiter else ","
        self.encoding = encoding
        self.ar = None
        self.analyses = None
        self.worksheet = worksheet if worksheet else 0
        self.infile = infile
        self.csv_data = None
        self.sample_id = None
        self.processed_samples_class = []
        mimetype, encoding = guess_type(self.infile.filename)
        InstrumentResultsFileParser.__init__(self, infile, mimetype)

    def xls_to_csv(self, infile, worksheet=0, delimiter=","):
        """
        Convert xlsx to easier format first, since we want to use the
        convenience of the CSV library
        """

        wb = open_workbook(file_contents=infile.read())
        sheet = wb.sheets()[worksheet]

        buffer = StringIO()

        # extract all rows
        for row in sheet.get_rows():
            line = []
            for cell in row:
                value = cell.value
                if type(value) in types.StringTypes:
                    value = value.encode("utf8")
                if value is None:
                    value = ""
                line.append(str(value))
            print >> buffer, delimiter.join(line)
        buffer.seek(0)
        return buffer

    def xlsx_to_csv(self, infile, worksheet=None, delimiter=","):
        worksheet = worksheet if worksheet else 0
        wb = load_workbook(filename=infile)
        if worksheet in wb.sheetnames:
            sheet = wb[worksheet]
        else:
            try:
                index = int(worksheet)
                sheet = wb.worksheets[index]
            except (ValueError, TypeError, IndexError):
                raise SheetNotFound

        buffer = StringIO()

        for row in sheet.rows:
            line = []
            for cell in row:
                new_val = ''
                if cell.number_format == "0.00%":
                    new_val = '{}%'.format(cell.value * 100)
                cellval = new_val if new_val else cell.value
                if (isinstance(cellval, (int, long, float))):
                    value = "" if cellval is None else str(cellval).encode(
                                    "utf8")
                else:
                    value = "" if cellval is None else cellval.encode(
                                    "utf8")
                if "\n" in value:
                    value = value.split("\n")[0]
                line.append(value.strip())
            if not any(line):
                continue
            buffer.write(delimiter.join(line) + "\n")
        buffer.seek(0)
        return buffer

    def parse(self):
        order = []
        ext = splitext(self.infile.filename.lower())[-1]
        if ext == ".xlsx":
            order = (self.xlsx_to_csv, self.xls_to_csv)
        elif ext == ".xls":
            order = (self.xls_to_csv, self.xlsx_to_csv)
        elif ext == ".csv" or ".prn":
            self.csv_data = self.infile
        if order:
            for importer in order:
                try:
                    self.csv_data = importer(
                        infile=self.infile,
                        worksheet=self.worksheet,
                        delimiter=self.delimiter,
                    )
                    break
                except SheetNotFound:
                    self.err(
                        "Sheet not found in workbook: %s" % self.worksheet)
                    return -1
                except Exception as e:
                    pass
            else:
                self.warn("Can't parse input file as XLS, XLSX, or CSV.")
                return -1

        stub = FileStub(file=self.csv_data, name=str(self.infile.filename))
        self.csv_data = FileUpload(stub)

        data = self.csv_data.read()
        lines = self.data_cleaning(data, ext)

        analysis_round = 0
        sample_service, lines = self.parse_headerlines(lines)

        for row_nr, row in enumerate(lines):
            if "Mthode:" in row[0] or "Method:" in row[0]:
                analysis_round = analysis_round + 1
            elif analysis_round > 0 and row[0]:

                self.parse_row(
                    row, sample_service[analysis_round-1],
                    analysis_round, row_nr)
        return 0

    def parse_row(self, row, sample_service, analysis_round, row_nr):
        sample_id = row[0]
        reading = row[1]

        if not sample_id or not reading:
            self.warn(
                "Data not entered correctly for '{}' with sample ID '{}'"
                " and result of '{}'".format(sample_service,
                                             sample_id, reading))
            return 0

        # Here we check whether this sample ID has been processed already
        if {sample_id: sample_service} in self.processed_samples_class:
            msg = (
                "Multiple results for Sample '{}' with sample service '{}'"
                " found. Not imported".format(sample_id, sample_service))
            raise MultipleAnalysesFound(msg)

        interim_keywords = self.get_interim_fields(sample_id)
        try:
            if self.is_sample(sample_id):
                ar = self.get_ar(sample_id)
                analysis = self.get_analysis(ar, sample_service)
            elif self.is_analysis_group_id(sample_id):
                analysis = self.get_duplicate_or_qc(sample_id, sample_service)
            else:
                sample_reference = self.get_reference_sample(
                        sample_id, sample_service)
                analysis = self.get_reference_sample_analysis(
                        sample_reference, sample_service)
            if not analysis:
                keyword = self.process_interims(
                        interim_keywords, sample_service, sample_id, reading)
                if not keyword:
                    # keyword = analysis.getKeyword  # Will throw error
                    self.warn("No Analysis found for Sample {0} and keyword"
                              " {1}. Results not imported".format(
                                                sample_id, sample_service))
                    return
                else:
                    return
        except Exception as e:
            self.warn(
                msg="Error getting analysis for '${s}/${kw}': ${e}",
                mapping={'s': sample_id, 'kw': sample_service, 'e': repr(e)},
                numline=row_nr, line=str(row))
            return

        if reading == "OVER":
            if analysis_round == 3:
                reading = 999999
            else:
                return
        self.processed_samples_class.append({sample_id: sample_service})
        self.parse_results(float(reading), keyword, sample_id)
        return

    def process_interims(self, interims, kw, sample_id, result):
        as_kw = interims.get(kw)
        if as_kw:
            if len(as_kw) > 1:
                self.warn("Duplicate keyword {0} found for Sample {1} and"
                          " their results are not imported".format(
                                                kw, sample_id))
                return "Duplicate"
            else:
                self.processed_samples_class.append({sample_id: kw})
                self.parse_interims(result, as_kw[0], kw, sample_id)
                return as_kw[0]
        else:
            return

    def parse_interims(self, result, as_kw, interim_kw, sample_id):
        parsed = {}
        parsed[interim_kw] = result
        parsed.update({"DefaultResult": interim_kw})
        self._addRawResult(sample_id, {as_kw: parsed})

    def parse_results(self, result, keyword, sample_id):
        parsed = {}
        parsed["Reading"] = float(result)
        parsed.update({"DefaultResult": "Reading"})
        self._addRawResult(sample_id, {keyword: parsed})

    @staticmethod
    def extract_relevant_data(lines):
        new_lines = []
        for row in lines:
            split_row = row.encode("ascii", "ignore").split(",")
            new_lines.append(split_row)
        return new_lines

    def data_cleaning(self, data, ext):
        decoded_data = self.try_utf8(data)
        if decoded_data:
            if ext == ".xlsx":
                lines_with_parentheses = decoded_data.split("\n")
            else:
                lines_with_parentheses = decoded_data.split("\r\n")
        else:
            decoded_data = self.try_utf16(data)
            if decoded_data:
                if "\r\n" in decoded_data:
                    lines_with_parentheses = data.decode(
                                                'utf-16').split("\r\n")
                elif "\n" in decoded_data:
                    lines_with_parentheses = data.decode(
                                                'utf-16').split("\n")
                else:
                    lines_with_parentheses = re.sub(
                            r'[^\x00-\x7f]',
                            r'', data).split("\r\n")  # if bad conversion
            else:
                if "\r\n" in data:
                    lines_with_parentheses = re.sub(
                            r'[^\x00-\x7f]', r'', data).split("\r\n")
                else:
                    lines_with_parentheses = re.sub(
                            r'[^\x00-\x7f]', r'', data).split("\n")
        lines = [i.replace('"', '') for i in lines_with_parentheses]

        ascii_lines = self.extract_relevant_data(lines)
        return ascii_lines

    @staticmethod
    def get_interim_fields(sample_id):
        bc = api.get_tool(CATALOG_ANALYSIS_REQUEST_LISTING)
        ar = bc(portal_type='AnalysisRequest', id=sample_id)
        if len(ar) == 0:
            ar = bc(
                portal_type='AnalysisRequest', getClientSampleID=sample_id)
        if len(ar) == 1:
            obj = ar[0].getObject()
            analyses = obj.getAnalyses(full_objects=True)
            keywords = {}
            for analysis_service in analyses:
                for field in analysis_service.getInterimFields():
                    interim_kw = field.get("keyword")
                    as_kw = analysis_service.Keyword
                    if interim_kw in keywords.keys():
                        keywords[interim_kw].append(as_kw)
                    else:
                        keywords[interim_kw] = [as_kw]
            return keywords
        return {}

    @staticmethod
    def parse_headerlines(lines):
        sample_service = []
        for row_nr, row in enumerate(lines):
            if row_nr == 5:
                return sample_service, lines[5:]
            if "Mthodes" in row[0] or "Methods" in row[0]:
                # Determining how many rounds there are in the sheet (Max = 3)
                if row[1]:
                    sample_service.append(row[1])
                if len(row) > 2 and row[2]:
                    sample_service.append(row[2])
                if len(row) > 3 and row[3]:
                    sample_service.append(row[3])

    @staticmethod
    def try_utf8(data):
        """Returns a Unicode object on success, or None on failure"""
        try:
            return data.decode('utf-8')
        except UnicodeDecodeError:
            return None

    @staticmethod
    def try_utf16(data):
        """Returns a Unicode object on success, or None on failure"""
        try:
            return data.decode('utf-16')
        except UnicodeDecodeError:
            return None

    @staticmethod
    def is_sample(sample_id):
        query = dict(portal_type="AnalysisRequest", getId=sample_id)
        brains = api.search(query, CATALOG_ANALYSIS_REQUEST_LISTING)
        return True if brains else False

    @staticmethod
    def get_ar(sample_id):
        query = dict(portal_type="AnalysisRequest", getId=sample_id)
        brains = api.search(query, CATALOG_ANALYSIS_REQUEST_LISTING)
        try:
            return api.get_object(brains[0])
        except IndexError:
            pass

    def get_analysis(self, ar, kw):
        analyses = self.get_analyses(ar)
        analyses = [v for k, v in analyses.items() if k == kw]
        if len(analyses) < 1:
            self.log(' No analysis found matching keyword {}'.format(kw))
            return None
        if len(analyses) > 1:
            self.warn('Multiple analyses found matching Keyword {}'.format(kw))
            return None
        return analyses[0]

    @staticmethod
    def get_analyses(ar):
        analyses = ar.getAnalyses()
        return dict((a.getKeyword, a) for a in analyses)

    @staticmethod
    def is_analysis_group_id(analysis_group_id):
        portal_types = ["DuplicateAnalysis", "ReferenceAnalysis"]
        query = dict(
            portal_type=portal_types,
            getReferenceAnalysesGroupID=analysis_group_id
        )
        brains = api.search(query, ANALYSIS_CATALOG)
        return True if brains else False

    @staticmethod
    def get_duplicate_or_qc(analysis_id, sample_service, ):
        portal_types = ["DuplicateAnalysis", "ReferenceAnalysis"]
        query = dict(
            portal_type=portal_types, getReferenceAnalysesGroupID=analysis_id
        )
        brains = api.search(query, ANALYSIS_CATALOG)
        analyses = dict((a.getKeyword, a) for a in brains)
        brains = [
                v for k, v in analyses.items() if k == sample_service]
        if len(brains) < 1:
            msg = (
                " No analysis found matching Keyword {}".format(
                                                            sample_service))
            raise AnalysisNotFound(msg)
        if len(brains) > 1:
            msg = (
                "Multiple brains found matching Keyword {}".format(
                                                            sample_service))
            raise MultipleAnalysesFound(msg)
        return brains[0]

    @staticmethod
    def get_reference_sample(reference_sample_id, kw):
        query = dict(
            portal_type="ReferenceSample", getId=reference_sample_id
        )
        brains = api.search(query, SENAITE_CATALOG)
        if len(brains) < 1:
            msg = ("No reference sample found matching Keyword {}".format(kw))
            raise AnalysisNotFound(msg)
        if len(brains) > 1:
            msg = ("Multiple brains found matching Keyword {}".format(kw))
            raise MultipleAnalysesFound(msg)
        return brains[0]

    @staticmethod
    def get_reference_sample_analyses(reference_sample):
        brains = reference_sample.getObject().getReferenceAnalyses()
        return dict((a.getKeyword(), a) for a in brains)

    def get_reference_sample_analysis(self, reference_sample, kw):
        kw = kw
        brains = self.get_reference_sample_analyses(reference_sample)
        brains = [v for k, v in brains.items() if k == kw]
        if len(brains) < 1:
            msg = " No analysis found matching Keyword {}".format(kw)
            raise AnalysisNotFound(msg)
        if len(brains) > 1:
            msg = ("Multiple brains found matching Keyword {}".format(kw))
            raise MultipleAnalysesFound(msg)
        return brains[0]


class flameatomicimport(object):
    implements(IInstrumentImportInterface, IInstrumentAutoImportInterface)
    title = "Agilent Flame Atomic Absorption"
    __file__ = abspath(__file__)  # noqa

    def __init__(self, context):
        self.context = context
        self.request = None

    @staticmethod
    def Import(context, request):
        errors = []
        logs = []
        warns = []

        infile = request.form["instrument_results_file"]
        if not hasattr(infile, "filename"):
            errors.append(_("No file selected"))

        artoapply = request.form["artoapply"]
        override = request.form["results_override"]
        instrument = request.form.get("instrument", None)
        worksheet = request.form.get("worksheet", 0)
        parser = FlameAtomicParser(infile, worksheet=worksheet)
        if parser:

            status = ["sample_received", "attachment_due", "to_be_verified"]
            if artoapply == "received":
                status = ["sample_received"]
            elif artoapply == "received_tobeverified":
                status = [
                        "sample_received", "attachment_due", "to_be_verified"]

            over = [False, False]
            if override == "nooverride":
                over = [False, False]
            elif override == "override":
                over = [True, False]
            elif override == "overrideempty":
                over = [True, True]

            importer = AnalysisResultsImporter(
                parser=parser,
                context=context,
                allowed_ar_states=status,
                allowed_analysis_states=None,
                override=over,
                instrument_uid=instrument,
            )

            try:
                importer.process()
                errors = importer.errors
                logs = importer.logs
                warns = importer.warns
            except Exception as e:
                errors.extend([repr(e), traceback.format_exc()])

        results = {"errors": errors, "log": logs, "warns": warns}

        return json.dumps(results)


class MyExport(BrowserView):

    def __innit__(self, context, request):
        self.context = context
        self.request = request

    def __call__(self, analyses):
        now = DateTime().strftime('%Y%m%d-%H%M')
        uc = api.get_tool('uid_catalog')
        instrument = self.context.getInstrument()
        norm = getUtility(IIDNormalizer).normalize
        filename = '{}-{}.lbl'.format(
            self.context.getId(), instrument.Title())

        options = {
            'dilute_factor': 1,
            'method': 'Dilution',
            'notneeded1': 1,
            'notneeded2': 1,
            'notneeded3': 10
        }

        sample_cases = {'a': 'SAMP', 'b': 'BLANK', 'c': 'CRM', 'd': 'DUP'}

        layout = self.context.getLayout()
        tmprows = []
        sample_dict = {}

        for indx, item in enumerate(layout):
            c_uid = item['container_uid']
            a_uid = item['analysis_uid']
            analysis = uc(UID=a_uid)[0].getObject() if a_uid else None
            keyword = str(analysis.Keyword)
            container = uc(UID=c_uid)[0].getObject() if c_uid else None
            sample_type = sample_cases[item['type']]
            kw_P = None

            if item['type'] == 'a':
                # sample_id
                analysis_id = container.id
                ans = container.getAnalyses()
                # PeseePourFusion keywords - kw_P
                kw_P = filter(lambda x: x.getKeyword == "PeseePourFusion", ans)
            elif (item['type'] in 'bcd'):
                analysis_id = analysis.getReferenceAnalysesGroupID()
            weight = 0
            if kw_P:
                peseepourfusion = kw_P[0]
                weight = peseepourfusion.getResult
            if not weight:
                weight = 0
            tmprows.append([indx+1,
                            analysis_id,
                            sample_type,
                            weight,
                            options['dilute_factor'],
                            options["notneeded1"],
                            options["notneeded2"],
                            options["notneeded3"]])

        for row in tmprows:
            if sample_dict.get(row[1]):
                sample_dict[row[1]].append(row)
            else:
                sample_dict[row[1]] = [row]

        unsorted_rows = []
        for rows in sample_dict.values():
            max_weight = -1
            for items in rows:
                if items[3] > max_weight:
                    max_weight = items[3]
            rows[0][3] = max_weight
            unsorted_rows.append(rows[0])

        unsorted_rows.sort(lambda a, b: cmp(a[0], b[0]))
        final_rows = self.row_sorter(unsorted_rows)
        result = self.dict_to_string(final_rows)

        setheader = self.request.RESPONSE.setHeader
        setheader('Content-Length', len(result))
        setheader('Content-Disposition', 'attachment; filename=%s' % filename)
        setheader('Content-Type', 'text/lbl')
        self.request.RESPONSE.write(result)

    @staticmethod
    def dict_to_string(rows):
        final_rows = ''
        interim_rows = []

        for row in rows:
            row = ','.join(str(item) for item in row)
            interim_rows.append(row)
        final_rows = '\r\n'.join(interim_rows)
        return final_rows

    @staticmethod
    def row_sorter(rows):
        for indx, row in enumerate(rows):
            row[0] = indx+1
        return rows


class flameatomicexport(object):
    implements(IInstrumentExportInterface)
    title = "Agilent Flame Atomic Exporter"
    __file__ = abspath(__file__)  # noqa

    def __init__(self, context, request=None):
        self.context = context
        self.request = request

    def Export(self, context, request):
        return MyExport(context, request)
