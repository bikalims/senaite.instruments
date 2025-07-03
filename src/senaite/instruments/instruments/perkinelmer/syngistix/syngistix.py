# -*- coding: utf-8 -*-
#
# This file is part of SENAITE.INSTRUMENTS.
#
# SENAITE.CORE is free software: you can redistribute it and/or modify it under
# the terms of the GNU General Public License as published by the Free Software
# Foundation, version 2.
#
# This program is distributed in the hope that it will be useful, but WITHOUT
# ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS
# FOR A PARTICULAR PURPOSE. See the GNU General Public License for more
# details.
#
# You should have received a copy of the GNU General Public License along with
# this program; if not, write to the Free Software Foundation, Inc., 51
# Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.
#
# Copyright 2018-2019 by it's authors.
# Some rights reserved, see README and LICENSE.

import csv
import json
import types
import traceback
from Products.CMFPlone.utils import safe_unicode
from cStringIO import StringIO
from mimetypes import guess_type
from openpyxl import load_workbook
from os.path import abspath
from os.path import splitext
from re import subn
from xlrd import open_workbook
from zope.interface import implements
from zope.publisher.browser import FileUpload

from bika.lims import api
from bika.lims.browser import BrowserView
from bika.lims.catalog import CATALOG_ANALYSIS_REQUEST_LISTING
from senaite.instruments import senaiteMessageFactory as _
from senaite.core.catalog import ANALYSIS_CATALOG, SENAITE_CATALOG
from senaite.core.exportimport.instruments import (
    IInstrumentAutoImportInterface,
    IInstrumentExportInterface,
    IInstrumentImportInterface,
)
from senaite.core.exportimport.instruments.resultsimport import (
    AnalysisResultsImporter,
    InstrumentResultsFileParser,
)
from senaite.instruments.instrument import FileStub
from senaite.instruments.instrument import SheetNotFound


class SampleNotFound(Exception):
    pass


class MultipleAnalysesFound(Exception):
    pass


class AnalysisNotFound(Exception):
    pass


class SyngistixParser(InstrumentResultsFileParser):
    ar = None

    def __init__(self, infile, worksheet=None, encoding=None, delimiter=None):
        self.delimiter = delimiter if delimiter else ","
        self.encoding = encoding
        self.ar = None
        self.analyses = None
        self.worksheet = worksheet if worksheet else 0
        self.infile = infile
        self.csv_data = None
        self.sample_id = None
        mimetype, encoding = guess_type(self.infile.filename)
        InstrumentResultsFileParser.__init__(self, infile, mimetype)

    def xls_to_csv(self, infile, worksheet="", delimiter=","):
        """
        Convert xlsx to easier format first, since we want to use the
        convenience of the CSV library

        """

        def find_sheet(wb, worksheet):
            for sheet in wb.sheets():
                if sheet.name == worksheet:
                    return sheet

        wb = open_workbook(file_contents=infile.read())
        sheet = wb.sheets()[worksheet]

        buffer = StringIO()

        # extract all rows
        for row in sheet.get_rows():
            line = []
            for cell in row:
                value = cell.value
                if type(value) in types.StringTypes:
                    value = value.encode("utf8")
                if value is None:
                    value = ""
                line.append(str(value))
            print >> buffer, delimiter.join(line)
        buffer.seek(0)
        return buffer

    def xlsx_to_csv(self, infile, worksheet="", delimiter=","):
        worksheet = worksheet if worksheet else 0
        wb = load_workbook(filename=infile)
        if worksheet in wb.sheetnames:
            sheet = wb[worksheet]
        else:
            try:
                index = int(worksheet)
                sheet = wb.worksheets[index]
            except (ValueError, TypeError, IndexError):
                raise SheetNotFound

        buffer = StringIO()
        for row in sheet.rows:
            line = []
            for cell in row:
                new_val = ""
                if cell.number_format == "0.00%":
                    new_val = "{}%".format(cell.value * 100)
                cellval = new_val if new_val else cell.value

                try:
                    value = (
                        "" if cellval is None else str(cellval).encode("utf8")
                    )
                except UnicodeEncodeError:
                    value = (
                        ""
                        if cellval is None
                        else safe_unicode(cellval).encode("utf8")
                    )
                if "\n" in value:  # fixme multi-line cell gives only 1st line
                    value = value.split("\n")[0]
                line.append(value.strip())
            if not any(line):
                continue
            buffer.write(delimiter.join(line) + "\n")
        buffer.seek(0)
        return buffer

    def parse(self):
        order = []
        ext = splitext(self.infile.filename.lower())[-1]
        if ext == ".xlsx":
            order = (self.xlsx_to_csv, self.xls_to_csv)
        elif ext == ".xls":
            order = (self.xls_to_csv, self.xlsx_to_csv)
        elif ext == ".csv":
            self.csv_data = self.infile
        if order:
            for importer in order:
                try:
                    self.csv_data = importer(
                        infile=self.infile,
                        worksheet=self.worksheet,
                        delimiter=self.delimiter,
                    )
                    break
                except SheetNotFound:
                    self.err(
                        "Sheet not found in workbook: %s" % self.worksheet
                    )
                    return -1
                except Exception as e:  # noqa
                    pass
            else:
                self.warn("Can't parse input file as XLS, XLSX, or CSV.")
                return -1
        stub = FileStub(file=self.csv_data, name=str(self.infile.filename))
        self.csv_data = FileUpload(stub)

        portal_type = ""
        lines = self.csv_data.readlines()
        reader = csv.DictReader(lines)
        results = []
        for row in reader:
            new_row = self.remove_unwanted_columns(row)
            results.append(new_row)

        row_num = 2
        for row in results:
            sample_id = row.get("Sample Id", "")
            del row["Sample Id"]
            portal_type = self.get_portal_type(sample_id)
            if portal_type == "AnalysisRequest":
                self.parse_ar_row(sample_id, row_num, row)
            elif portal_type in ["DuplicateAnalysis", "ReferenceAnalysis"]:
                self.parse_duplicate_row(sample_id, row_num, row)

            elif portal_type == "ReferenceSample":
                self.parse_reference_sample_row(sample_id, row_num, row)
            else:
                self.warn(
                    msg="No results found for '${sample_id}'",
                    mapping={"sample_id": sample_id},
                    numline=str(reader.line_num),
                )
            row_num = row_num + 1
        return 1

    def get_portal_type(self, sample_id):
        portal_type = None
        if self.is_sample(sample_id):
            ar = self.get_ar(sample_id)
            self.ar = ar
            self.analyses = self.get_analyses(ar)
            portal_type = ar.portal_type
        elif self.is_analysis_group_id(sample_id):
            portal_type = "DuplicateAnalysis"
        elif self.is_reference_sample(sample_id):
            portal_type = "ReferenceSample"
        return portal_type

    def parse_row(self, row_nr, parsed, sample_id):
        #  interim_kw = "Reading"
        #  parsed.update({"DefaultResult": interim_kw})
        self._addRawResult(sample_id, parsed)
        return 0

    def parse_ar_row(self, sample_id, row_nr, row):
        ar = self.get_ar(sample_id)
        items = row.items()
        edited_items = {k.split(" ", 1)[0]: v for k, v in items if k}
        items = edited_items.items()
        interim_kw = "Reading"
        parsed = {
            subn(r"[^\w\d\-_]*", "", k)[0]: {interim_kw: v}
            for k, v in items
            if k
        }
        for item in items:
            keyword = item[0]
            try:
                analysis = self.get_analysis(ar, keyword)
                analysis_obj = analysis.getObject()
                import pdb; pdb.set_trace()
                analysis.InterimFields
                interim_fields = analysis_obj.InterimFields
                precision = analysis_obj.Precision
                field_kws = [x.get("keyword") for x in interim_fields if x]
                if "Reading" not in field_kws:
                    self.warn(
                        msg="No interim field 'Reading' was found for Analysis '${kw}' on ${sample_id}. Result was not imported.",
                        mapping={"kw": keyword, "sample_id": sample_id},
                        numline=row_nr,
                    )
                    del parsed[keyword]
                else:
                    result = float(parsed[keyword][interim_kw])
                    rounded_result = round(result, precision)
                    parsed[keyword][interim_kw] = str(rounded_result)
                if not analysis:
                    del parsed[keyword]
            except Exception:
                self.warn(
                    msg="Error getting analysis for '${kw}': ${sample_id}",
                    mapping={"kw": keyword, "sample_id": sample_id},
                    numline=row_nr,
                )
                del parsed[keyword]
        return self.parse_row(row_nr, parsed, sample_id)

    def parse_duplicate_row(self, sample_id, row_nr, row):
        items = row.items()
        edited_items = {k.split(" ", 1)[0]: v for k, v in items if k}
        items = edited_items.items()
        interim_kw = "Reading"
        parsed = {
            subn(r"[^\w\d\-_]*", "", k)[0]: {interim_kw: v}
            for k, v in items
            if k
        }
        for item in items:
            keyword = item[0]
            try:
                analysis = self.get_duplicate_or_qc_analysis(sample_id, keyword)
                Dup_keyword = self.getDuplicateKeyword(analysis)
                precision = analysis.getObject().Precision
                if not Dup_keyword:
                    del parsed[keyword]
                elif "No Interim Field" in Dup_keyword:
                    self.warn(
                        msg="No interim field 'Reading' was found for Analysis '${kw}' on ${sample_id}. Result was not imported.",
                        mapping={"kw": keyword, "sample_id": sample_id},
                        numline=row_nr,
                    )
                    del parsed[keyword]
                else:
                    result = float(parsed[keyword][interim_kw])
                    rounded_result = round(result, precision)
                    parsed[keyword][interim_kw] = str(rounded_result)
            except Exception:
                self.warn(
                    msg="Error getting analysis for '${kw}': ${sample_id}",
                    mapping={"kw": keyword, "sample_id": sample_id},
                    numline=row_nr,
                )
                del parsed[keyword]
        return self.parse_row(row_nr, parsed, sample_id)

    def getDuplicateKeyword(self, analysis):
        keyword = analysis.getKeyword
        if analysis:
            interim_fields = analysis.getObject().InterimFields
            field_kws = [x.get("keyword") for x in interim_fields if x]
            if "Reading" not in field_kws:
                keyword = "No Interim Field"
        return keyword

    def parse_reference_sample_row(self, sample_id, row_nr, row):
        items = row.items()
        parsed = {subn(r"[^\w\d\-_]*", "", k)[0]: v for k, v in items if k}

        keyword = "DU_SCC"
        try:
            if not self.getReferenceSampleKeyword(sample_id, keyword):
                return 0
        except Exception:
            self.warn(
                msg="Error getting analysis for '${kw}': ${sample_id}",
                mapping={"kw": keyword, "sample_id": sample_id},
                numline=row_nr,
            )
            return
        return self.parse_row(row_nr, parsed, keyword)

    def getReferenceSampleKeyword(self, sample_id, kw):
        sample_reference = self.get_reference_sample(sample_id, kw)
        analysis = self.get_reference_sample_analysis(sample_reference, kw)
        return analysis.getKeyword()

    def get_reference_sample_analysis(self, reference_sample, kw):
        kw = kw
        brains = self.get_reference_sample_analyses(reference_sample)
        brains = [v for k, v in brains.items() if k == kw]
        if len(brains) < 1:
            lmsg = "No analysis found for sample {} matching Keyword {}"
            msg = lmsg.format(reference_sample, kw)
            raise AnalysisNotFound(msg)
        if len(brains) > 1:
            lmsg = "Multiple objects found for sample {} matching Keyword '{}'"
            msg = lmsg.format(reference_sample, kw)
            raise MultipleAnalysesFound(msg)
        return brains[0]

    @staticmethod
    def get_reference_sample_analyses(reference_sample):
        brains = reference_sample.getObject().getReferenceAnalyses()
        return dict((a.getKeyword(), a) for a in brains)

    @staticmethod
    def get_duplicate_or_qc_analysis(analysis_id, kw):
        portal_types = ["DuplicateAnalysis", "ReferenceAnalysis"]
        query = dict(
            portal_type=portal_types, getReferenceAnalysesGroupID=analysis_id
        )
        brains = api.search(query, ANALYSIS_CATALOG)
        analyses = dict((a.getKeyword, a) for a in brains)
        brains = [v for k, v in analyses.items() if k == kw]
        if len(brains) < 1:
            lmsg = "No analysis found for sample {} matching Keyword {}"
            msg = lmsg.format(analysis_id, kw)
            raise AnalysisNotFound(msg)
        if len(brains) > 1:
            lmsg = "Multiple objects found for sample {} matching Keyword {}"
            msg = lmsg.format(analysis_id, kw)
            raise MultipleAnalysesFound(msg)
        return brains[0]

    @staticmethod
    def get_ar(sample_id):
        query = dict(portal_type="AnalysisRequest", getId=sample_id)
        brains = api.search(query, CATALOG_ANALYSIS_REQUEST_LISTING)
        try:
            return api.get_object(brains[0])
        except IndexError:
            pass

    @staticmethod
    def is_sample(sample_id):
        query = dict(portal_type="AnalysisRequest", getId=sample_id)
        brains = api.search(query, CATALOG_ANALYSIS_REQUEST_LISTING)
        return True if brains else False

    @staticmethod
    def get_analyses(ar):
        analyses = ar.getAnalyses()
        return dict((a.getKeyword, a) for a in analyses)

    def get_analysis(self, ar, kw):
        analyses = self.get_analyses(ar)
        analyses = [v for k, v in analyses.items() if k == kw]
        if len(analyses) < 1:
            msg = (
                "No analysis found for sample '${ar}' matching keyword '${kw}'"
            )
            self.log(msg, mapping=dict(kw=kw, ar=ar.getId()))
            return None
        if len(analyses) > 1:
            self.warn(
                'Multiple analyses found matching Keyword "${kw}"',
                mapping=dict(kw=kw),
            )
            return None
        return analyses[0]

    @staticmethod
    def is_analysis_group_id(analysis_group_id):
        portal_types = ["DuplicateAnalysis", "ReferenceAnalysis"]
        query = dict(
            portal_type=portal_types,
            getReferenceAnalysesGroupID=analysis_group_id,
        )
        brains = api.search(query, ANALYSIS_CATALOG)
        return True if brains else False

    @staticmethod
    def is_reference_sample(reference_sample_id):
        query = dict(portal_type="ReferenceSample", getId=reference_sample_id)
        brains = api.search(query, SENAITE_CATALOG)
        return True if brains else False

    @staticmethod
    def get_reference_sample(reference_sample_id, kw):
        query = dict(portal_type="ReferenceSample", getId=reference_sample_id)
        brains = api.search(query, SENAITE_CATALOG)
        if len(brains) < 1:
            lmsg = (
                "No reference sample found for sample {} matching Keyword {}"
            )
            msg = lmsg.format(reference_sample_id, kw)
            raise AnalysisNotFound(msg)
        if len(brains) > 1:
            lmsg = "Multiple objects found for sample {} matching Keyword {}"
            msg = lmsg.format(reference_sample_id, kw)
            raise MultipleAnalysesFound(msg)
        return brains[0]

    def remove_unwanted_columns(self, row):
        del row["R"]
        del row["Acquisition Time"]
        del row["A/S Loc"]
        del row["QC Status"]
        del row["Dataset File"]
        del row["Method File"]
        del row[""]
        return row


class importer(object):
    implements(IInstrumentImportInterface, IInstrumentAutoImportInterface)
    title = "Perkin Elmer Syngistix"
    __file__ = abspath(__file__)  # noqa

    def __init__(self, context):
        self.context = context
        self.request = None

    @staticmethod
    def Import(context, request):
        errors = []
        logs = []
        warns = []

        infile = request.form["instrument_results_file"]
        if not hasattr(infile, "filename"):
            errors.append(_("No file selected"))

        artoapply = request.form["artoapply"]
        override = request.form["results_override"]
        instrument = request.form.get("instrument", None)
        worksheet = "Conc. in Sample Units"  # The required worksheet's name
        parser = SyngistixParser(infile, worksheet=worksheet)
        if parser:

            status = ["sample_received", "attachment_due", "to_be_verified"]
            if artoapply == "received":
                status = ["sample_received"]
            elif artoapply == "received_tobeverified":
                status = [
                    "sample_received",
                    "attachment_due",
                    "to_be_verified",
                ]

            over = [False, False]
            if override == "nooverride":
                over = [False, False]
            elif override == "override":
                over = [True, False]
            elif override == "overrideempty":
                over = [True, True]

            importer = AnalysisResultsImporter(
                parser=parser,
                context=context,
                allowed_sample_states=status,
                allowed_analysis_states=None,
                override=over,
                instrument_uid=instrument,
            )

            try:
                importer.process()
                errors = importer.errors
                logs = importer.logs
                warns = importer.warns
            except Exception as e:
                errors.extend([repr(e), traceback.format_exc()])

        results = {"errors": errors, "log": logs, "warns": warns}

        return json.dumps(results)


class MyExport(BrowserView):
    def __init__(self, context, request):
        self.context = context
        self.request = request

    def __call__(self, analyses):
        uc = api.get_tool("uid_catalog")
        instrument = self.context.getInstrument()
        filename = "{}-{}.csv".format(self.context.getId(), instrument.Title())

        layout = self.context.getLayout()
        tmprows = []
        parsed_analyses = {}
        headers = ["Sample Label", "Weight", "Volume", "Dilution"]
        tmprows.append(headers)

        for indx, item in enumerate(layout):
            c_uid = item["container_uid"]
            a_uid = item["analysis_uid"]
            analysis = uc(UID=a_uid)[0].getObject() if a_uid else None
            container = uc(UID=c_uid)[0].getObject() if c_uid else None

            if item["type"] == "a":
                analysis_id = container.id
            elif item["type"] in "bcd":
                analysis_id = analysis.getReferenceAnalysesGroupID()
            if parsed_analyses.get(analysis_id):
                continue
            else:
                tmprows.append([analysis_id, "", "", ""])
                parsed_analyses[analysis_id] = 10

        result = self.dict_to_string(tmprows)

        setheader = self.request.RESPONSE.setHeader
        setheader("Content-Length", len(result))
        setheader("Content-Disposition", "inline; filename=%s" % filename)
        setheader("Content-Type", "text/csv")
        self.request.RESPONSE.write(result)

    @staticmethod
    def dict_to_string(rows):
        final_rows = ""
        interim_rows = []

        for row in rows:
            row = ",".join(str(item) for item in row)
            interim_rows.append(row)
        final_rows = "\r\n".join(interim_rows)
        return final_rows


class syngistixesxport(object):
    implements(IInstrumentExportInterface)
    title = "Perkin Elmer Syngistix Exporter"
    __file__ = abspath(__file__)  # noqa

    def __init__(self, context, request=None):
        self.context = context
        self.request = request

    def Export(self, context, request):
        return MyExport(context, request)
