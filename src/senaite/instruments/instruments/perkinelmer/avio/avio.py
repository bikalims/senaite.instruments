# -*- coding: utf-8 -*-
"""Importer for Perkin Elmer Syngistix Uranium result workbooks."""

import json
import re
import traceback
from mimetypes import guess_type
from os.path import abspath
from os.path import splitext

from openpyxl import load_workbook
from zope.interface import implements

from senaite.instruments import senaiteMessageFactory as _
from senaite.core.exportimport.instruments import (
    IInstrumentAutoImportInterface,
    IInstrumentImportInterface,
)
from senaite.core.exportimport.instruments.resultsimport import (
    AnalysisResultsImporter,
)
from senaite.instruments.instruments.pg.dv5000icp.dv5000 import (
    DV5000ICPParser,
)


WORKSHEET = "Corrected Intensities"


class AvioParser(DV5000ICPParser):
    """Read results exclusively from the Corrected Intensities sheet."""

    def __init__(self, infile, worksheet=WORKSHEET, encoding=None):
        # The sheet argument is accepted for importer API compatibility, but
        # this instrument must never read concentration or RSD sheets.
        super(AvioParser, self).__init__(
            infile, worksheet=WORKSHEET, encoding=encoding)

    def parse(self):
        ext = splitext(self.infile.filename.lower())[-1]
        if ext not in (".xlsx", ".xlsm"):
            self.err("Unsupported file format: %s" % ext)
            return -1

        try:
            workbook = load_workbook(filename=self.infile, data_only=True)
        except Exception:
            self.err("Cannot open workbook: %s" % self.infile.filename)
            return -1

        try:
            worksheet = workbook[WORKSHEET]
        except KeyError:
            self.err("Sheet not found in workbook: %s" % WORKSHEET)
            return -1

        # Syngistix repeats the header whenever the analyte set changes.
        headers = []
        for row_nr in range(1, worksheet.max_row + 1):
            sample_id = self.safe_value(
                worksheet.cell(row=row_nr, column=2).value)
            if sample_id == "Sample Id":
                headers = self.get_result_headers(worksheet, row_nr)
                continue
            if not sample_id or not headers:
                continue

            values = [
                self.safe_value(
                    worksheet.cell(row=row_nr, column=column).value)
                for column, header in headers
            ]
            self.headers = [header for column, header in headers]
            self.parse_result_row(sample_id, row_nr, values)

        return 1

    @staticmethod
    def get_result_headers(worksheet, row_nr):
        headers = []
        # Columns A-G contain run metadata. Results start in column H.
        for column in range(8, worksheet.max_column + 1):
            value = AvioParser.safe_value(
                worksheet.cell(row=row_nr, column=column).value)
            if value:
                headers.append((column, value))
        return headers

    @staticmethod
    def normalize_keyword(header):
        """Turn ``Fe-lb 238.863 (cps)`` into ``Fe238863``."""
        header = re.sub(r"\b(?:lb|cps)\b", "", header,
                        flags=re.IGNORECASE)
        return re.sub(r"[^A-Za-z0-9]", "", header)


class importer(object):
    implements(IInstrumentImportInterface, IInstrumentAutoImportInterface)
    title = "Perkin Elmer Syngistix Uranium"
    __file__ = abspath(__file__)  # noqa

    def __init__(self, context):
        self.context = context
        self.request = None

    @staticmethod
    def Import(context, request):
        errors = []
        logs = []
        warns = []

        infile = request.form["instrument_results_file"]
        if not hasattr(infile, "filename"):
            errors.append(_("No file selected"))

        artoapply = request.form["artoapply"]
        override = request.form["results_override"]
        instrument = request.form.get("instrument", None)
        parser = AvioParser(infile)

        status = ["sample_received", "attachment_due", "to_be_verified"]
        if artoapply == "received":
            status = ["sample_received"]

        over = [False, False]
        if override == "override":
            over = [True, False]
        elif override == "overrideempty":
            over = [True, True]

        results_importer = AnalysisResultsImporter(
            parser=parser,
            context=context,
            allowed_sample_states=status,
            allowed_analysis_states=None,
            override=over,
            instrument_uid=instrument,
        )
        try:
            results_importer.process()
            errors = results_importer.errors
            logs = results_importer.logs
            warns = results_importer.warns
        except Exception as error:
            errors.extend([repr(error), traceback.format_exc()])

        return json.dumps({"errors": errors, "log": logs, "warns": warns})

    def get_automatic_importer(self, instrument, parser, **kw):
        if getattr(parser, "_instrument", None) is None:
            parser._instrument = instrument
        self.parser = parser
        return self

    def get_automatic_parser(self, infile):
        return AvioParser(infile)

    def process(self):
        return self.Import(self.context, request=None)

