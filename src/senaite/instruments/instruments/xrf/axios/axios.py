# -*- coding: utf-8 -*-
"""Importer for Malvern Panalytical Axios XRF result files."""

import csv
import json
import re
import traceback
from cStringIO import StringIO
from mimetypes import guess_type
from os.path import abspath
from os.path import splitext

from openpyxl import load_workbook
from xlrd import open_workbook
from zope.interface import implements

from bika.lims import api
from bika.lims.catalog import CATALOG_ANALYSIS_REQUEST_LISTING
from senaite.core.catalog import ANALYSIS_CATALOG
from senaite.core.exportimport.instruments import IInstrumentAutoImportInterface
from senaite.core.exportimport.instruments import IInstrumentImportInterface
from senaite.core.exportimport.instruments.resultsimport import AnalysisResultsImporter
from senaite.core.exportimport.instruments.resultsimport import InstrumentResultsFileParser
from senaite.instruments import senaiteMessageFactory as _


IDENTITY_HEADERS = ("sample name", "seq", "meas date/time", "sum",
                    "result type")
INTERIM_HEADERS = {
    "seq": "Seq",
    "meas date/time": "MeasDateTime",
    "sum": "SumOfConc",
    "result type": "ResultType",
}


class AxiosXRFParser(InstrumentResultsFileParser):
    """Parse the result table in an Axios CSV, XLS, or XLSX export."""

    def __init__(self, infile, worksheet=0, encoding=None, delimiter=None):
        self.infile = infile
        self.worksheet = worksheet if worksheet is not None else 0
        self.encoding = encoding or "utf-8"
        self.delimiter = delimiter
        mimetype, unused = guess_type(infile.filename)
        InstrumentResultsFileParser.__init__(self, infile, mimetype)

    def parse(self):
        try:
            rows = self.read_rows()
        except Exception:
            self.err("Unable to read '{}':\n{}".format(
                self.infile.filename, traceback.format_exc()))
            return -1

        header_nr, columns, result_columns = self.find_header(rows)
        if header_nr is None:
            self.err("Could not find the 'Sample name' result table header")
            return -1
        if not result_columns:
            self.err("No analyte result columns were found")
            return -1

        for row_nr, row in enumerate(rows[header_nr + 1:], header_nr + 2):
            sample_id = self.cell(row, columns["sample name"])
            if not sample_id:
                continue
            analyses = self.get_analyses(sample_id)
            if not analyses:
                self.warn(
                    msg="Sample or QC ID '${sample_id}' was not found",
                    mapping={"sample_id": sample_id}, numline=row_nr)
                continue

            parsed = {}
            for col_nr, instrument_keyword in result_columns:
                value = self.cell(row, col_nr)
                if value == "":
                    continue
                if not self.is_number(value):
                    self.warn(
                        msg="Invalid result '${result}' for '${sample_id}'",
                        mapping={"result": value, "sample_id": sample_id},
                        numline=row_nr)
                    continue
                self.add_result(parsed, analyses, instrument_keyword, value,
                                row, columns, sample_id, row_nr)

            if parsed:
                self._addRawResult(sample_id, parsed)
        return 1

    def add_result(self, parsed, analyses, instrument_keyword, value, row,
                   columns, sample_id, row_nr):
        """Add a direct result, or use the header as an interim keyword."""
        direct = [a for a in analyses
                  if self.analysis_keyword(a) == instrument_keyword]
        if len(direct) == 1:
            payload = {"Result": value, "DefaultResult": "Result"}
            payload.update(self.get_interims(direct[0], row, columns))
            parsed[instrument_keyword] = payload
            return

        interim = [a for a in analyses
                   if instrument_keyword in self.interim_keywords(a)]
        if len(interim) == 1:
            keyword = self.analysis_keyword(interim[0])
            payload = {instrument_keyword: value,
                       "DefaultResult": instrument_keyword}
            payload.update(self.get_interims(interim[0], row, columns))
            parsed[keyword] = payload
            return

        reason = "multiple analyses match" if len(direct + interim) > 1 \
            else "no analysis or interim field matches"
        self.warn(
            msg=("Result '${kw}' for '${sample_id}' was not imported: "
                 "${reason}"),
            mapping={"kw": instrument_keyword, "sample_id": sample_id,
                     "reason": reason}, numline=row_nr)

    def get_interims(self, analysis, row, columns):
        configured = self.interim_keywords(analysis)
        values = {}
        for header, interim in INTERIM_HEADERS.items():
            if interim in configured and header in columns:
                values[interim] = self.cell(row, columns[header])
        return values

    @staticmethod
    def analysis_keyword(analysis):
        return analysis.getKeyword

    @staticmethod
    def interim_keywords(analysis):
        obj = analysis.getObject() if hasattr(analysis, "getObject") else analysis
        fields = obj.getInterimFields() if hasattr(obj, "getInterimFields") \
            else getattr(obj, "InterimFields", [])
        return [field.get("keyword") for field in fields if field]

    def get_analyses(self, sample_id):
        ar = self.get_ar(sample_id)
        if ar:
            return ar.getAnalyses()
        query = dict(
            portal_type=["DuplicateAnalysis", "ReferenceAnalysis"],
            getReferenceAnalysesGroupID=sample_id,
        )
        return api.search(query, ANALYSIS_CATALOG)

    @staticmethod
    def get_ar(sample_id):
        query = dict(portal_type="AnalysisRequest", getId=sample_id)
        brains = api.search(query, CATALOG_ANALYSIS_REQUEST_LISTING)
        return api.get_object(brains[0]) if brains else None

    def find_header(self, rows):
        for row_nr, row in enumerate(rows):
            normalized = [self.normalize_label(value) for value in row]
            sample_columns = [i for i, value in enumerate(normalized)
                              if value.startswith("sample name")]
            if not sample_columns:
                continue
            columns = {"sample name": sample_columns[0]}
            result_columns = []
            for col_nr, value in enumerate(normalized):
                identity = self.identity_header(value)
                if identity:
                    columns[identity] = col_nr
                elif value:
                    result_columns.append(
                        (col_nr, self.normalize_keyword(row[col_nr])))
            return row_nr, columns, result_columns
        return None, {}, []

    @staticmethod
    def identity_header(value):
        for header in IDENTITY_HEADERS:
            if value == header or value.startswith(header):
                return header
        return None

    def read_rows(self):
        extension = splitext(self.infile.filename.lower())[-1]
        self.infile.seek(0)
        if extension == ".csv":
            data = self.infile.read()
            if isinstance(data, unicode):
                data = data.encode(self.encoding)
            delimiter = self.delimiter or self.detect_delimiter(data)
            return [self.clean_row(row) for row in
                    csv.reader(StringIO(data), delimiter=delimiter)]
        if extension in (".xlsx", ".xlsm"):
            workbook = load_workbook(filename=self.infile, data_only=True)
            sheet = self.get_xlsx_sheet(workbook)
            return [[self.safe_value(cell.value) for cell in row]
                    for row in sheet.iter_rows()]
        if extension == ".xls":
            workbook = open_workbook(file_contents=self.infile.read())
            sheet = workbook.sheet_by_name(self.worksheet) \
                if isinstance(self.worksheet, basestring) \
                else workbook.sheet_by_index(int(self.worksheet))
            return [[self.safe_value(sheet.cell_value(r, c))
                    for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        raise ValueError("Unsupported file format: {}".format(extension))

    def get_xlsx_sheet(self, workbook):
        if isinstance(self.worksheet, basestring):
            if self.worksheet not in workbook.sheetnames:
                raise ValueError("Sheet not found: {}".format(self.worksheet))
            return workbook[self.worksheet]
        return workbook.worksheets[int(self.worksheet)]

    @staticmethod
    def detect_delimiter(data):
        try:
            return csv.Sniffer().sniff(data[:4096], ",;\t|").delimiter
        except csv.Error:
            return ","

    @staticmethod
    def safe_value(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return unicode(value).strip()

    @classmethod
    def clean_row(cls, row):
        return [cls.safe_value(value) for value in row]

    @classmethod
    def normalize_label(cls, value):
        return " ".join(cls.safe_value(value).lower().replace(".", "").split())

    @classmethod
    def normalize_keyword(cls, value):
        return re.sub(r"[^\w\d\-_]*", "", cls.safe_value(value))

    @staticmethod
    def cell(row, col_nr):
        return AxiosXRFParser.safe_value(row[col_nr]) \
            if col_nr is not None and col_nr < len(row) else ""

    @staticmethod
    def is_number(value):
        try:
            float(value)
            return True
        except (TypeError, ValueError):
            return False


class importer(object):
    implements(IInstrumentImportInterface, IInstrumentAutoImportInterface)
    title = "Malvern Panalytical Axios XRF"
    __file__ = abspath(__file__)  # noqa

    def __init__(self, context):
        self.context = context
        self.parser = None

    @staticmethod
    def Import(context, request):
        errors, logs, warns = [], [], []
        infile = request.form["instrument_results_file"]
        if not hasattr(infile, "filename"):
            return json.dumps({"errors": [_('No file selected')],
                               "log": logs, "warns": warns})
        parser = AxiosXRFParser(
            infile, worksheet=request.form.get("worksheet", 0))
        states = ["sample_received", "attachment_due", "to_be_verified"]
        if request.form.get("artoapply") == "received":
            states = ["sample_received"]
        override = request.form.get("results_override", "overrideempty")
        over = {"override": [True, False],
                "overrideempty": [True, True]}.get(override, [False, False])
        results_importer = AnalysisResultsImporter(
            parser=parser, context=context, allowed_sample_states=states,
            allowed_analysis_states=None, override=over,
            instrument_uid=request.form.get("instrument"))
        try:
            results_importer.process()
            errors.extend(results_importer.errors)
            logs.extend(results_importer.logs)
            warns.extend(results_importer.warns)
        except Exception:
            errors.append(traceback.format_exc())
        return json.dumps({"errors": errors, "log": logs, "warns": warns})

    def get_automatic_parser(self, infile):
        return AxiosXRFParser(infile)

    def get_automatic_importer(self, instrument, parser, **kwargs):
        self.parser = parser
        return self

    def process(self):
        request = type("Request", (object,), {})()
        request.form = dict(
            instrument_results_file=self.parser.infile,
            artoapply="received_tobeverified", results_override="overrideempty",
            instrument=None, worksheet=self.parser.worksheet)
        return self.Import(self.context, request)
