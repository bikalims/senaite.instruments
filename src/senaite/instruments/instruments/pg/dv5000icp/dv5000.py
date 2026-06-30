# -*- coding: utf-8 -*-
"""SENAITE instrument importer for PG DV5000 ICP result workbooks.

Expected workbook layout, as supplied by the instrument:

    Sheet:  Result
    Row 1:  Analyte headers, e.g. Ca396.847-A, Fe259.940-A, ...
    Row 2:  Units, e.g. ppm
    Blocks: Sample ID + acquisition timestamp in column A
            replicate rows 1, 2, 3...
            mean row marked with the greek chi character: χ
            sigma row: σ
            RSD row: RSD%

Only the mean row (χ) is imported.
"""

import json
import re
import traceback
from mimetypes import guess_type
from os.path import abspath
from os.path import splitext

from openpyxl import load_workbook
from zope.interface import implements

from bika.lims import api
from bika.lims.catalog import CATALOG_ANALYSIS_REQUEST_LISTING
from senaite.instruments import senaiteMessageFactory as _
from senaite.core.catalog import ANALYSIS_CATALOG
from senaite.core.exportimport.instruments import (
    IInstrumentAutoImportInterface,
    IInstrumentImportInterface,
)
from senaite.core.exportimport.instruments.resultsimport import (
    AnalysisResultsImporter,
    InstrumentResultsFileParser,
)
from senaite.instruments.instrument import SheetNotFound


MEAN_MARKERS = (u"χ", "x", "X", "mean", "Mean", "MEAN")

# Override here if SENAITE analysis keywords do not match the instrument headers.
# Example:
# HEADER_KEYWORD_MAP = {
#     "Ca396.847-A": "Ca",
#     "Fe259.940-A": "Fe",
# }
HEADER_KEYWORD_MAP = {}


class SampleNotFound(Exception):
    pass


class MultipleAnalysesFound(Exception):
    pass


class AnalysisNotFound(Exception):
    pass


class DV5000ICPParser(InstrumentResultsFileParser):
    """Parser for PG DV5000 ICP Excel result files."""

    def __init__(self, infile, worksheet="Result", encoding=None):
        self.infile = infile
        self.worksheet = worksheet or "Result"
        self.encoding = encoding
        self.headers = []
        self.units = []
        self.ar = None
        self.analyses = None
        mimetype, enc = guess_type(self.infile.filename)
        InstrumentResultsFileParser.__init__(self, infile, mimetype)

    def parse(self):
        ext = splitext(self.infile.filename.lower())[-1]
        if ext not in (".xlsx", ".xlsm"):
            self.err("Unsupported file format: %s" % ext)
            return -1

        try:
            wb = load_workbook(filename=self.infile, data_only=True)
        except Exception:
            self.err("Cannot open workbook: %s" % self.infile.filename)
            return -1

        try:
            ws = wb[self.worksheet]
        except KeyError:
            self.err("Sheet not found in workbook: %s" % self.worksheet)
            return -1

        self.headers = self.get_headers(ws)
        self.units = self.get_units(ws)
        if not self.headers:
            self.err("No analyte headers found in row 1")
            return -1

        sample_id = None
        for row_nr in range(3, ws.max_row + 1):
            marker = self.safe_value(ws.cell(row=row_nr, column=1).value)
            if not marker:
                continue

            if self.is_sample_header(marker):
                sample_id = self.extract_sample_id(marker)
                continue

            if sample_id and marker in MEAN_MARKERS:
                values = [
                    self.safe_value(ws.cell(row=row_nr, column=col_nr).value)
                    for col_nr in range(2, len(self.headers) + 2)
                ]
                self.parse_result_row(sample_id, row_nr, values)

        return 1

    def get_headers(self, ws):
        headers = []
        for col_nr in range(2, ws.max_column + 1):
            value = self.safe_value(ws.cell(row=1, column=col_nr).value)
            if value:
                headers.append(value)
        return headers

    def get_units(self, ws):
        units = []
        for col_nr in range(2, len(self.headers) + 2):
            units.append(self.safe_value(ws.cell(row=2, column=col_nr).value))
        return units

    @staticmethod
    def safe_value(value):
        if value is None:
            return ""
        try:
            return unicode(value).strip()
        except NameError:
            return str(value).strip()

    @staticmethod
    def is_sample_header(value):
        """A sample header is the first row of a sample/result block."""
        if value in MEAN_MARKERS:
            return False
        if value in (u"σ", "RSD%"):
            return False
        if value.isdigit():
            return False
        if value.startswith("InstrumentCode:"):
            return False
        return True

    @staticmethod
    def extract_sample_id(value):
        """Remove the trailing acquisition timestamp from column A."""
        value = value.strip()
        value = re.sub(
            r"\s+\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2}:\d{2}$",
            "",
            value,
        )
        return value.strip()

    @staticmethod
    def normalize_keyword(header):
        """Convert instrument header to a SENAITE-friendly keyword.

        By default the dot is removed and hyphen is preserved:
        Ca396.847-A -> Ca396847-A

        Use HEADER_KEYWORD_MAP above when the lab service keyword differs.
        """
        if header in HEADER_KEYWORD_MAP:
            return HEADER_KEYWORD_MAP[header]
        return re.sub(r"[^\w\d\-_]*", "", header)

    def parse_result_row(self, sample_id, row_nr, values):
        portal_type = self.get_portal_type(sample_id)
        if portal_type == "AnalysisRequest":
            return self.parse_ar_row(sample_id, row_nr, values)
        if portal_type in ["DuplicateAnalysis", "ReferenceAnalysis"]:
            return self.parse_duplicate_and_reference_row(sample_id, row_nr, values)
        self.warn(
            msg="No results found for '${sample_id}'",
            mapping={"sample_id": sample_id},
            numline=row_nr,
        )
        return 0

    def parse_ar_row(self, sample_id, row_nr, values):
        ar = self.get_ar(sample_id)
        if not ar:
            self.warn(
                msg="Sample '${sample_id}' was not found",
                mapping={"sample_id": sample_id},
                numline=row_nr,
            )
            return 0

        parsed = {}
        for header, value in zip(self.headers, values):
            keyword = self.normalize_keyword(header)
            if not keyword or value == "":
                continue
            analysis = self.get_analysis(ar, keyword)
            if not analysis:
                continue
            parsed[keyword] = value

        self._addRawResult(sample_id, parsed)
        return 0

    def parse_duplicate_and_reference_row(self, sample_id, row_nr, values):
        parsed = {}
        for header, value in zip(self.headers, values):
            keyword = self.normalize_keyword(header)
            if not keyword or value == "":
                continue
            try:
                analysis = self.get_duplicate_or_qc_analysis(sample_id, keyword)
            except Exception:
                self.warn(
                    msg="Error getting analysis for '${kw}': ${sample_id}",
                    mapping={"kw": keyword, "sample_id": sample_id},
                    numline=row_nr,
                )
                continue
            if analysis:
                parsed[keyword] = value

        self._addRawResult(sample_id, parsed)
        return 0

    def get_portal_type(self, sample_id):
        if self.is_sample(sample_id):
            ar = self.get_ar(sample_id)
            self.ar = ar
            self.analyses = self.get_analyses(ar)
            return ar.portal_type
        if self.is_analysis_group_id(sample_id):
            return "DuplicateAnalysis"
        return None

    @staticmethod
    def get_ar(sample_id):
        query = dict(portal_type="AnalysisRequest", getId=sample_id)
        brains = api.search(query, CATALOG_ANALYSIS_REQUEST_LISTING)
        try:
            return api.get_object(brains[0])
        except IndexError:
            return None

    @staticmethod
    def is_sample(sample_id):
        query = dict(portal_type="AnalysisRequest", getId=sample_id)
        brains = api.search(query, CATALOG_ANALYSIS_REQUEST_LISTING)
        return True if brains else False

    @staticmethod
    def get_analyses(ar):
        analyses = ar.getAnalyses()
        return dict((a.getKeyword, a) for a in analyses)

    def get_analysis(self, ar, kw):
        analyses = self.get_analyses(ar)
        analyses = [v for k, v in analyses.items() if k == kw]
        if len(analyses) < 1:
            self.warn(
                msg="No analysis found for sample '${ar}' matching keyword '${kw}'",
                mapping=dict(kw=kw, ar=ar.getId()),
            )
            return None
        if len(analyses) > 1:
            self.warn(
                'Multiple analyses found matching Keyword "${kw}"',
                mapping=dict(kw=kw),
            )
            return None
        return analyses[0]

    @staticmethod
    def is_analysis_group_id(analysis_group_id):
        portal_types = ["DuplicateAnalysis", "ReferenceAnalysis"]
        query = dict(
            portal_type=portal_types,
            getReferenceAnalysesGroupID=analysis_group_id,
        )
        brains = api.search(query, ANALYSIS_CATALOG)
        return True if brains else False

    @staticmethod
    def get_duplicate_or_qc_analysis(analysis_id, kw):
        portal_types = ["DuplicateAnalysis", "ReferenceAnalysis"]
        query = dict(
            portal_type=portal_types,
            getReferenceAnalysesGroupID=analysis_id,
        )
        brains = api.search(query, ANALYSIS_CATALOG)
        analyses = dict((a.getKeyword, a) for a in brains)
        brains = [v for k, v in analyses.items() if k == kw]
        if len(brains) < 1:
            msg = "No analysis found for sample {} matching Keyword {}"
            raise AnalysisNotFound(msg.format(analysis_id, kw))
        if len(brains) > 1:
            msg = "Multiple objects found for sample {} matching Keyword {}"
            raise MultipleAnalysesFound(msg.format(analysis_id, kw))
        return brains[0]


class importer(object):
    implements(IInstrumentImportInterface, IInstrumentAutoImportInterface)
    title = "PG DV5000 ICP"
    __file__ = abspath(__file__)  # noqa

    def __init__(self, context):
        self.context = context
        self.request = None

    @staticmethod
    def Import(context, request):
        errors = []
        logs = []
        warns = []

        infile = request.form["instrument_results_file"]
        if not hasattr(infile, "filename"):
            errors.append(_("No file selected"))

        artoapply = request.form["artoapply"]
        override = request.form["results_override"]
        instrument = request.form.get("instrument", None)
        worksheet = request.form.get("worksheet", "Result")
        parser = DV5000ICPParser(infile, worksheet=worksheet)

        if parser:
            status = ["sample_received", "attachment_due", "to_be_verified"]
            if artoapply == "received":
                status = ["sample_received"]
            elif artoapply == "received_tobeverified":
                status = [
                    "sample_received",
                    "attachment_due",
                    "to_be_verified",
                ]

            over = [False, False]
            if override == "nooverride":
                over = [False, False]
            elif override == "override":
                over = [True, False]
            elif override == "overrideempty":
                over = [True, True]

            results_importer = AnalysisResultsImporter(
                parser=parser,
                context=context,
                allowed_sample_states=status,
                allowed_analysis_states=None,
                override=over,
                instrument_uid=instrument,
            )

            try:
                results_importer.process()
                errors = results_importer.errors
                logs = results_importer.logs
                warns = results_importer.warns
            except Exception as e:
                errors.extend([repr(e), traceback.format_exc()])

        results = {"errors": errors, "log": logs, "warns": warns}
        return json.dumps(results)

    def get_automatic_importer(self, instrument, parser, **kw):
        if getattr(parser, "_instrument", None) is None:
            parser._instrument = instrument
        self.parser = parser
        return self

    def get_automatic_parser(self, infile):
        return DV5000ICPParser(infile, worksheet="Result")

    def process(self):
        return self.Import(self.context, request=None)
