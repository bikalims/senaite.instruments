import types

from openpyxl import load_workbook
from senaite.core.exportimport.instruments.resultsimport import \
    InstrumentResultsFileParser
from cStringIO import StringIO
from xlrd import open_workbook
from zope.publisher.browser import FileUpload


def xls_to_csv(infile, worksheet=0, delimiter=","):
    # TODO: Move to utility module
    """
    Convert xlsx to easier format first, since we want to use the
    convenience of the CSV library

    """

    def find_sheet(wb, worksheet):
        for sheet in wb.sheets():
            if sheet.name == worksheet:
                return sheet

    wb = open_workbook(file_contents=infile.read())
    sheet = wb.sheets()[worksheet]

    buffer = StringIO()

    # extract all rows
    for row in sheet.get_rows():
        line = []
        for cell in row:
            value = cell.value
            if type(value) in types.StringTypes:
                value = value.encode("utf8")
            if value is None:
                value = ""
            line.append(str(value))
        print >> buffer, delimiter.join(line)  # noqa
    buffer.seek(0)
    return buffer


class SheetNotFound(Exception):
    """
    Sheet not found in workbook
    """


def xlsx_to_csv(infile, worksheet=0, delimiter=",", data_only=True):
    wb = load_workbook(filename=infile, data_only=data_only)

    if isinstance(worksheet, int):
        if worksheet >= len(wb.sheetnames):
            return
        worksheet = wb.sheetnames[worksheet]
    if worksheet not in wb.sheetnames:
        return

    sheet = wb[worksheet]

    buffer = StringIO()
    for row in sheet.rows:
        line = []
        for cell in row:
            value = "" if cell.value is None else str(cell.value).encode("utf8")  # noqa
            if "\n" in value:  # fixme multi-line cell gives only first line
                value = value.split("\n")[0]
            line.append(value.strip())
        if not any(line):
            continue
        buffer.write(delimiter.join(line) + "\n")
    buffer.seek(0)
    return buffer


class FileStub:

    def __init__(self, file, name):
        self.file = file
        self.headers = {}
        self.filename = name


class InstrumentXLSResultsFileParser(InstrumentResultsFileParser):
    """ Parser
    """

    def __init__(self, infile, worksheet=0, encoding='xlsx', delimiter=None, data_only=False): # noqa
        InstrumentResultsFileParser.__init__(self, infile, encoding.upper())
        # Convert xls to csv
        self._delimiter = delimiter if delimiter else "|"
        self._worksheet = worksheet
        if encoding == 'xlsx':
            csv_data = xlsx_to_csv(
                infile,
                worksheet=worksheet,
                delimiter=self._delimiter,
                data_only=data_only)
        elif encoding == 'xls':
            csv_data = xls_to_csv(
                infile, worksheet=worksheet, delimiter=self._delimiter)

        if not csv_data:
            self._infile = None
            self._csvfile = None
            self.err("Could not load worksheet {}".format(worksheet)) # noqa
            return

        # adapt csv_data into a FileUpload for parse method
        self._infile = infile
        stub = FileStub(file=csv_data, name=str(infile.filename))
        self._csvfile = FileUpload(stub)

        self._encoding = encoding
        self._end_header = False

    def parse(self):
        if self._csvfile is None:
            return True

        infile = self._csvfile
        self.log("Parsing worksheet '${worksheet}' of file '${file_name}'",
                 mapping={
                    "worksheet": self._worksheet,
                    "file_name": infile.filename
                 })
        jump = 0
        # We test in import functions if the file was uploaded
        try:
            name = getattr(infile, 'filename', getattr(infile, 'name'))
            f = open(name, 'rU')
        except AttributeError:
            f = infile

        for line in f.readlines():
            self._numline += 1
            if jump == -1:
                # Something went wrong. Finish
                self.err("File processing finished due to critical errors")
                return False
            if jump > 0:
                # Jump some lines
                jump -= 1
                continue

            if not line or not line.strip():
                continue

            line = line.strip()
            jump = 0
            if line:
                jump = self._parseline(line)

        self.log(
            "End of file reached successfully: ${total_objects} objects, "
            "${total_analyses} analyses, ${total_results} results",
            mapping={"total_objects": self.getObjectsTotalCount(),
                     "total_analyses": self.getAnalysesTotalCount(),
                     "total_results": self.getResultsTotalCount()}
        )
        return True
